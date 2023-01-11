# Data Formatting
from openpyxl import Workbook, load_workbook

def data_format():
    v5_data = load_workbook('v5_data_clean.xlsx')
    ws_v5_12 = v5_data['v5_12']
    ws_v5_11 = v5_data['v5_11']

    formatted_data = Workbook()
    v5_12_sheet = formatted_data.active
    v5_12_sheet.title = 'v5_12'
    formatted_data.create_sheet('v5_11')
    v5_11_sheet = formatted_data['v5_11']

    v5_12_sheet.append(['Sl No', 'Number','TID', 'Running', 'Closed'])
    v5_11_sheet.append(['Sl No', 'Number','TID', 'Running', 'Closed'])

    exg1_number = 2222274000  # Exchange 1 starting number
    for row in range (2, 1600+5):
        v5_12_sheet['A' + str(row)].value = row-1   # Serial Number make
        v5_12_sheet['C' + str(row)].value = row-1 # TID Make
        v5_12_sheet['B' + str(row)].value = '0' + str(exg1_number)  # Number Generate
        exg1_number = exg1_number + 1

    formatted_data.save('formatted_data.xlsx')

    source_number = 2
    # Running and close number separate
    for row in range (2, 1600+5):
        number = v5_12_sheet['B' + str(row)].value
        running_number = ws_v5_12['B' + str(source_number)].value
        # Running number finding
        if(number == running_number):
            v5_12_sheet['D' + str(row)].value = number
            source_number = source_number + 1
            #print('Run: ' + number)
        #Close number finding
        else:
            v5_12_sheet['E' + str(row)].value = number



    exg2_number = 2222275600  # Exchange 2 starting number
    for row in range (2, 800+5):
        v5_11_sheet['A' + str(row)].value = row-1   # Serial Number make
        v5_11_sheet['C' + str(row)].value = row-1 # TID Make
        v5_11_sheet['B' + str(row)].value = '0' + str(exg2_number)  # Number Generate
        exg2_number = exg2_number + 1

    formatted_data.save('formatted_data.xlsx')

    source_number = 2
    # Running and close number separate
    for row in range (2, 800+5):
        number = v5_11_sheet['B' + str(row)].value
        running_number = ws_v5_11['B' + str(source_number)].value

        # Running number finding
        if(number == running_number):
            v5_11_sheet['D' + str(row)].value = number
            source_number = source_number + 1
            #print('Run: ' + number)
        #Close number finding
        else:
            v5_11_sheet['E' + str(row)].value = number

    formatted_data.save('formatted_data.xlsx')