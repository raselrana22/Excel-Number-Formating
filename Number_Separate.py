# Data Cleaning
import openpyxl as xl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

def data_clean():
    v5_data = xl.load_workbook('V5_data.xlsx')

    ws_v5_12 = v5_data['v5-12']
    ws_v5_11 = v5_data['v5-11']


    v5_data_clean = Workbook()
    v5_12_sheet = v5_data_clean.active
    v5_12_sheet.title = "v5_12"
    v5_11_sheet = v5_data_clean.create_sheet('v5_11')
    v5_12_sheet.append(['Sl No', 'Number', 'TID'])
    v5_11_sheet.append(['Sl No', 'Number', 'TID'])

    new_number = ''
    for row in range (3, ws_v5_12.max_row-2):
        for col in range (1, 4):
            sl_no = row-2
            number = ws_v5_12.cell(row, 4).value
            tid = ws_v5_12.cell(row, 3).value

            char = get_column_letter(col)
            if char == 'A':
                v5_12_sheet[char + str(row-1)].value = sl_no
            elif char == 'B':
                for i in range(7, 18):
                    new_number += number[i]

                v5_12_sheet[char + str(row-1)].value = new_number
                new_number = ''   # Reset the variable
            elif char == 'C':
                v5_12_sheet[char + str(row-1)].value = tid

    for row in range (5, ws_v5_11.max_row-3):
       for col in range (1, 4):
            sl_no = row
            number = ws_v5_11.cell(row, 4).value
            tid = ws_v5_11.cell(row, 3).value

            char = get_column_letter(col)
            if char == 'A':
                v5_11_sheet[char + str(row-3)].value = sl_no
            elif char == 'B':
                for i in range(7, 18):
                    new_number += number[i]

                v5_11_sheet[char + str(row-3)].value = new_number
                new_number = ''   # Reset the variable
            elif char == 'C':
                v5_11_sheet[char + str(row-3)].value = tid

    v5_data_clean.save("v5_data_clean.xlsx")